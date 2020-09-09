from openpyxl import *
from tkinter import *
import tkinter as tk  
from functools import partial 

wb = load_workbook(r'C:\Users\HP\Desktop\db1.xlsx')
sheet = wb.active
def excel():
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 50
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.cell(row=1,column=1).value = "number1"
    sheet.cell(row=1,column=2).value = "number2"
    sheet.cell(row=1,column=3).value = "number3"
    sheet.cell(row=1,column=4).value = "result"
def focus1(event):
    no1_field.focus_set()
def focus2(event):
    no2_field.focus_set()
def focus3(event):
    no3_field.focus_set()
def focus4(event):
    no4_field.focus_set()

def clear():
    no1_field.delete(0,END)
    no2_field.delete(0,END)
    no3_field.delete(0,END)

def insert():
        if (no1_field.get()=="" and no2_field.get()=="" and no3_field.get()==""):
            print("empty input")
        else:
            current_row = sheet.max_row
            current_column = sheet.max_column
        
        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = email_field.get()
        sheet.cell(row=current_row + 1, column=3).value = password_field.get()

wb.save(r'C:\Users\HP\Desktop\db.xlsx')

    
    
    
    
def call_result(label_result, n1, n2, n3):  
    num1 = (n1.get())  
    num2 = (n2.get())  
    num3 = (n3.get())
    result = int(num1)+int(num2)+int(num3) 
    label_result.config(text="Result = %d" % result) 
    return  
def call_result1(label_result, n1, n2):  
    num1 = (n1.get())  
    num2 = (n2.get())  
    num3 = (n3.get())
    result = int(num1)-int(num2) 
    label_result.config(text="Result = %d" % result) 
    return
def call_result2(label_result, n1, n2, n3):  
    num1 = (n1.get())  
    num2 = (n2.get())  
    num3 = (n3.get())
    result = int(num1)*int(num2)*int(num3) 
    label_result.config(text="Result = %d" % result)
    return
   
root = tk.Tk()  
root.geometry('400x200+100+200')  
  
root.title('Calculator')  
   
number1 = tk.StringVar()  
number2 = tk.StringVar() 
number3 = tk.StringVar()
  
labelNum1 = tk.Label(root, text="A").grid(row=1, column=0)  
  
labelNum2 = tk.Label(root, text="B").grid(row=2, column=0) 

labelNum3 = tk.Label(root, text="C").grid(row=3, column=0)
  
labelResult = tk.Label(root)  
  
labelResult.grid(row=7, column=2)  
  
entryNum1 = tk.Entry(root, textvariable=number1).grid(row=1, column=2)  
  
entryNum2 = tk.Entry(root, textvariable=number2).grid(row=2, column=2)  

entryNum3 = tk.Entry(root, textvariable=number3).grid(row=3, column=2) 
  
call_result = partial(call_result, labelResult, number1, number2, number3)
call_result1 = partial(call_result1, labelResult, number1, number2) 
call_result2 = partial(call_result2, labelResult, number1, number2, number3) 
  
buttonCal = tk.Button(root, text="SUM", command=call_result).grid(row=5, column=1) 
buttonCal = tk.Button(root, text="SUBTRACT", command=call_result1).grid(row=5, column=2) 
buttonCal = tk.Button(root, text="MULTIPLY", command=call_result2).grid(row=5, column=3) 
  
root.mainloop()  
